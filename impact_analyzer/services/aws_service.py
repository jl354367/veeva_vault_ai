"""
AWS integration layer — the only file that talks to AWS.

Three responsibilities:
  1. S3 uploads — write the user's Data Model Changes file into S3.
  2. S3 reads   — grab the latest Config Report Excel from S3 (fallback path;
                  normally the Bedrock agent's action group fetches it).
  3. Bedrock    — send the user's chat message to the Bedrock agent and
                  return the agent's raw text response.

Layout of this file:
  • Small helpers (_boto_client, _is_aws_configured, _parse_excel_to_text).
  • Public S3 functions (fetch_config_report_from_s3, upload_document_to_s3).
  • Scope-hint helpers (_SCOPE_PHRASES, _detect_scope_hint) — used by
    invoke_agent to add a "STRICT SCOPE" directive when the user asks for
    a specific category (e.g. "workflows only").
  • invoke_agent — the star of the file. Wraps boto3's Bedrock call.
  • JSON extraction helpers (try_extract_report_json,
    _normalize_to_new_schema, _parse_agent_json, _repair_truncated_json)
    used by the router to pull structured data out of the agent's text.
"""

from __future__ import annotations

import io
import json
import logging
import re
from typing import Any

from config import settings

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# LOW-LEVEL HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _boto_client(service: str):
    """
    Build a boto3 client for the given service (e.g. "s3", "bedrock-agent-runtime").

    - Read timeout is 300s because Bedrock Agent responses can take 2-3 minutes
      when the agent invokes its own Lambda action groups.
    - boto3 is imported inside the function so the app can start even
      when boto3 isn't installed (e.g. during local UI-only development).
    - When AWS credentials aren't in .env, boto3 falls back to the default
      credential chain (IAM role, ~/.aws/credentials, env vars).
    """
    import boto3
    from botocore.config import Config
    # Bedrock Agent responses (with action groups) can take 2-3 minutes
    timeout_config = Config(read_timeout=300, connect_timeout=10, retries={"max_attempts": 0})
    kwargs: dict[str, Any] = {"region_name": settings.aws_region, "config": timeout_config}
    if settings.aws_access_key_id:
        kwargs["aws_access_key_id"] = settings.aws_access_key_id
        kwargs["aws_secret_access_key"] = settings.aws_secret_access_key
    return boto3.client(service, **kwargs)


def _is_aws_configured() -> bool:
    """True if we have enough config to talk to S3."""
    return bool(settings.s3_config_bucket and settings.aws_region)


# ─── Excel parser ─────────────────────────────────────────────────────────────
# Only used by the S3 fallback path — the Bedrock agent's own action group
# reads Excel files inside AWS Lambda, so this rarely runs in production.

def _parse_excel_to_text(excel_bytes: bytes) -> str:
    """
    Flatten an Excel workbook into a plain-text string, one line per row,
    columns joined with ' | '. Concatenates all sheets. Used when we need
    to feed Excel content to an LLM as raw text.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════
# S3 — download config report, upload user documents
# ═══════════════════════════════════════════════════════════════════════════

async def fetch_config_report_from_s3(
    bucket: str | None = None,
    key: str | None = None,
) -> dict[str, Any]:
    """Download the latest Vault Configuration Report Excel from S3."""
    bucket = bucket or settings.s3_config_bucket
    prefix = settings.s3_vault_reports_prefix

    if not _is_aws_configured():
        logger.warning("AWS not configured — returning mock config report")
        return _mock_config_report()

    try:
        s3 = _boto_client("s3")
        response = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)
        contents = response.get("Contents", [])
        excel_files = [
            obj for obj in contents
            if obj["Key"].lower().endswith((".xlsx", ".xlsm"))
        ]
        if not excel_files:
            logger.warning(
                "No Excel files found at s3://%s/%s — returning mock config report",
                bucket, prefix,
            )
            return _mock_config_report()
        latest = max(excel_files, key=lambda o: o["LastModified"])
        s3_key = latest["Key"]
        logger.info("Fetching latest config report: s3://%s/%s", bucket, s3_key)
        obj = s3.get_object(Bucket=bucket, Key=s3_key)
        excel_bytes = obj["Body"].read()
        report_text = _parse_excel_to_text(excel_bytes)
        logger.info("Parsed config report — %d chars from s3://%s/%s", len(report_text), bucket, s3_key)
        return {
            "source": "s3",
            "s3_key": s3_key,
            "report_text": report_text,
            "objects": [],
            "integrations": [],
        }
    except Exception as exc:
        logger.error("Failed to fetch config report from S3: %s", exc)
        raise


async def upload_document_to_s3(
    content: bytes,
    session_id: str,
    filename: str,
) -> str | None:
    """Upload a user-submitted document to the uploads bucket."""
    bucket = settings.s3_uploads_bucket
    if not bucket:
        logger.warning("S3_UPLOADS_BUCKET not configured — skipping document upload")
        return None
    import mimetypes
    s3_key = f"{settings.s3_uploads_prefix}{session_id}/{filename}"
    content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    s3 = _boto_client("s3")
    s3.put_object(Bucket=bucket, Key=s3_key, Body=content, ContentType=content_type)
    uri = f"s3://{bucket}/{s3_key}"
    logger.info("Uploaded document to %s", uri)
    return uri


# ═══════════════════════════════════════════════════════════════════════════
# SCOPE DETECTION
# ═══════════════════════════════════════════════════════════════════════════
# When the user asks a scoped question (e.g. "workflows only" or
# "list of common objects"), we detect the requested category and pass
# it as a STRICT SCOPE directive alongside the user's message.
# _SCOPE_PHRASES catches direct phrasings; _detect_scope_hint below
# also uses regex patterns for natural phrasings like
# "share the workflows" or "list of common integrations".

_SCOPE_PHRASES: tuple[tuple[str, str], ...] = (
    ("objects and fields", "objects and fields"),
    ("fields and objects", "objects and fields"),
    ("objects only", "objects"),
    ("just objects", "objects"),
    ("only objects", "objects"),
    ("fields only", "fields"),
    ("just fields", "fields"),
    ("only fields", "fields"),
    ("workflows only", "workflows"),
    ("just workflows", "workflows"),
    ("only workflows", "workflows"),
    ("integrations only", "integrations"),
    ("just integrations", "integrations"),
    ("only integrations", "integrations"),
    ("layouts only", "layouts"),
    ("just layouts", "layouts"),
    ("only layouts", "layouts"),
    ("picklists only", "picklists"),
    ("just picklists", "picklists"),
    ("only picklists", "picklists"),
    ("reports only", "reports"),
    ("just reports", "reports"),
    ("only reports", "reports"),
    ("sdk jobs only", "SDK jobs"),
    ("just sdk", "SDK jobs"),
    ("only sdk", "SDK jobs"),
    ("lifecycles only", "lifecycles"),
    ("validation only", "validation rules"),
    ("security only", "security and roles"),
)


def _detect_scope_hint(message: str) -> str | None:
    """
    Return a short scope description if the user's message asks for a
    specific subset of components (e.g. "objects and fields"), else None.

    Catches three families of phrasings:
      1. Explicit phrases from _SCOPE_PHRASES (e.g. "workflows only").
      2. "objects AND fields" / "fields AND objects".
      3. "(list|share|give|show) ... CATEGORY" where CATEGORY is one of
         objects/fields/workflows/integrations/layouts/picklists/reports/
         SDK jobs/lifecycles/validation rules/security — and the word is
         NOT followed by " and " (so combo asks fall through to #2).
      4. "common CATEGORY" — e.g. "common objects".
    """
    import re as _re
    low = message.lower()

    # 1) Direct phrase matches first
    for phrase, scope_desc in _SCOPE_PHRASES:
        if phrase in low:
            return scope_desc

    # 2) "objects AND fields" patterns — both categories at once
    if _re.search(
        r'(which|what|list|show|give\s+me|share|provide)\s+(?:the\s+)?(?:list\s+of\s+)?(?:common\s+)?objects?\s+and\s+fields?',
        low,
    ):
        return "objects and fields"
    if _re.search(
        r'(which|what|list|show|give\s+me|share|provide)\s+(?:the\s+)?(?:list\s+of\s+)?(?:common\s+)?fields?\s+and\s+objects?',
        low,
    ):
        return "objects and fields"

    # Reusable building blocks for the broader patterns below
    _action = r'(?:list|share|give|show|provide|tell\s+me)'
    _optional_prefix = (
        r'(?:\s+me)?'                        # "list me"
        r'(?:\s+(?:me\s+)?the)?'             # "list the", "list me the"
        r'(?:\s+(?:a\s+)?list\s+of)?'        # "list of"
        r'(?:\s+(?:all|common|every|only))?' # "all", "common", "every", "only"
        r'(?:\s+the)?'                       # extra "the"
    )

    _categories = (
        (r'objects?',                'objects'),
        (r'fields?',                 'fields'),
        (r'workflows?',              'workflows'),
        (r'lifecycles?',             'lifecycles'),
        (r'integrations?',           'integrations'),
        (r'(?:page\s+)?layouts?',    'layouts'),
        (r'picklists?',              'picklists'),
        (r'reports?',                'reports'),
        (r'sdk(?:\s+jobs?)?',        'SDK jobs'),
        (r'validation(?:\s+rules?)?','validation rules'),
        (r'security(?:\s+settings?)?','security and roles'),
    )

    # 3) "(list|share|give|show) ... CATEGORY" — singular scope
    for cat_re, scope_desc in _categories:
        pattern = _action + _optional_prefix + r'\s+' + cat_re + r'\b(?!\s+and\s+)'
        if _re.search(pattern, low):
            return scope_desc

    # 4) "common CATEGORY" anywhere — e.g. "what are the common objects?"
    for cat_re, scope_desc in _categories:
        if _re.search(r'\bcommon\s+' + cat_re + r'\b(?!\s+and\s+)', low):
            return scope_desc

    return None


# ═══════════════════════════════════════════════════════════════════════════
# BEDROCK AGENT — main entry point used by /stage1/chat
# ═══════════════════════════════════════════════════════════════════════════

async def invoke_agent(
    session_id: str,
    user_message: str,
    bedrock_session_id: str | None = None,
) -> dict[str, Any]:
    """
    Invoke the Bedrock Agent with the user's raw message.

    The agent (Claude) handles intent recognition, tool calls, and output
    format itself. The bedrock_session_id is reused across turns so the
    agent retains conversation memory and doesn't re-fetch documents.

    Returns:
        {
          "status":              "ok" | "error" | "not_configured",
          "response":            <raw text from agent in whatever format Claude chose>,
          "bedrock_session_id":  <session id to pass on next turn>,
          "error":               <error string when status == error>,
        }
    """
    if not settings.aws_agent_id:
        return {
            "status": "not_configured",
            "response": "Bedrock Agent is not configured. Set AWS_AGENT_ID and AWS_AGENT_ALIAS_ID in .env.",
        }

    import uuid as _uuid
    bedrock = _boto_client("bedrock-agent-runtime")
    bedrock_session_id = bedrock_session_id or str(_uuid.uuid4())

    # Send the user's message verbatim — same as the AWS Bedrock console
    # does. The only nudge we add is an emphatic SCOPE hint when the user
    # asks for a narrow subset (e.g. "objects only", "workflows only",
    # "common objects", "share the fields"); Nova Pro otherwise drifts
    # into listing every category in the report.
    scope_hint = _detect_scope_hint(user_message)
    if scope_hint:
        # All categories — used to spell out what NOT to include.
        all_cats = {
            "objects", "fields", "workflows", "lifecycles", "integrations",
            "layouts", "picklists", "reports", "SDK jobs",
            "validation rules", "security and roles",
        }
        excluded = sorted(
            c for c in all_cats
            if c.lower() not in scope_hint.lower()
        )
        input_text = (
            f"{user_message}\n\n"
            f"IMPORTANT — STRICT SCOPE: the user is asking for {scope_hint} ONLY. "
            f"Your response MUST contain ONLY {scope_hint}. "
            f"Do NOT include any of these other categories: "
            f"{', '.join(excluded)}. "
            f"If you start writing a section header for one of those, "
            f"stop immediately and remove it."
        )
    else:
        input_text = user_message

    logger.info(
        "Invoking Bedrock Agent agentId=%s aliasId=%s session=%s",
        settings.aws_agent_id, settings.aws_agent_alias_id, bedrock_session_id,
    )

    try:
        response = bedrock.invoke_agent(
            agentId=settings.aws_agent_id,
            agentAliasId=settings.aws_agent_alias_id,
            sessionId=bedrock_session_id,
            inputText=input_text,
        )
        agent_text = _collect_bedrock_stream(response)
        logger.info("Agent responded — %d chars", len(agent_text))
        return {
            "status": "ok",
            "response": agent_text,
            "bedrock_session_id": bedrock_session_id,
        }
    except Exception as exc:
        logger.error("Bedrock Agent invocation failed: %s", exc)
        return {
            "status": "error",
            "response": f"Sorry, I hit an error talking to the agent: {exc}",
            "error": str(exc),
            "bedrock_session_id": bedrock_session_id,
        }


# ═══════════════════════════════════════════════════════════════════════════
# JSON EXTRACTION — pull structured report data out of the agent's text
# ═══════════════════════════════════════════════════════════════════════════
# The agent normally answers in prose, but when the user asks for
# "as JSON" / "export" it also embeds a big JSON object. These helpers
# find that JSON, parse it, repair truncation if needed, and normalise
# legacy field names → new schema so the frontend always sees one shape.

def try_extract_report_json(text: str) -> dict[str, Any] | None:
    """
    Best-effort extraction of a Vault impact report JSON from the agent's
    response. Returns None if no valid impacted_objects JSON is found.

    The agent may respond in prose, table, or JSON. When it includes a JSON
    block (because the user asked for an export or by default), we extract
    it here so the frontend can render the structured report panel.
    """
    if not text or "impacted_objects" not in text:
        return None
    result = _parse_agent_json(text)
    if result:
        result = _normalize_to_new_schema(result)
    return result


def _normalize_to_new_schema(report: dict) -> dict:
    """
    Convert old-shape agent response to new shape if needed.
    Handles the case where the agent returns legacy keys
    (impacted_fields, severity, recommendations) instead of the
    new schema (impacted_areas, risk, recommendation).
    """
    if not isinstance(report, dict):
        return report

    for obj in report.get("impacted_objects", []) or []:
        # If already new shape with populated areas, skip
        if obj.get("impacted_areas"):
            continue

        areas = []
        old_fields = obj.get("impacted_fields", []) or []

        for f in old_fields:
            risk = (f.get("severity") or obj.get("overall_severity") or "MEDIUM").upper()
            old_def = f.get("old_definition", "") or ""
            new_def = f.get("new_definition", "") or ""
            change_desc = f.get("change_type", "") or ""
            desc = f.get("notes") or f"{change_desc}: {old_def} → {new_def}".strip(": → ")

            recs = obj.get("recommendations", []) or []
            recommendation = "; ".join(recs) if recs else "Review impacted field."

            areas.append({
                "area_type": "field",
                "area_name": f.get("field_name", ""),
                "description": desc,
                "risk": risk,
                "recommendation": recommendation,
            })

        # Old-shape integrations
        for i in (obj.get("impacted_integrations", []) or []):
            recs = obj.get("recommendations", []) or []
            areas.append({
                "area_type": "integration",
                "area_name": i.get("integration_name", ""),
                "description": i.get("notes") or f"Integration impact on {i.get('integration_name', '')}",
                "risk": (i.get("severity") or obj.get("overall_severity") or "MEDIUM").upper(),
                "recommendation": "; ".join(recs) if recs else "Review impacted integration.",
            })

        # No impacted_fields/integrations but object-level description exists →
        # create one area from object-level info
        if not areas and obj.get("description"):
            recs = obj.get("recommendations", []) or []
            areas.append({
                "area_type": obj.get("object_type", "other"),
                "area_name": obj.get("object_name", ""),
                "description": obj.get("description", ""),
                "risk": (obj.get("overall_severity") or "MEDIUM").upper(),
                "recommendation": "; ".join(recs) if recs else "Review change.",
            })

        obj["impacted_areas"] = areas

        # Promote first impacted_field to top-level field_name
        if old_fields and not obj.get("field_name"):
            obj["field_name"] = old_fields[0].get("field_name", "")
            obj["change_type"] = obj.get("change_type") or old_fields[0].get("change_type", "")

    return report


# ─── Streaming + JSON extraction helpers ─────────────────────────────────────

def _collect_bedrock_stream(response: dict) -> str:
    """Read all chunks from a Bedrock Agent streaming response into one string."""
    chunks = []
    for event in response.get("completion", []):
        chunk = event.get("chunk", {})
        if "bytes" in chunk:
            chunks.append(chunk["bytes"].decode("utf-8"))
    return "".join(chunks)


def _parse_agent_json(text: str) -> dict[str, Any] | None:
    """
    Extract JSON from agent response. Handles:
      - Prose preamble before JSON
      - Markdown code fences (```json ... ```)
      - Truncated JSON (token-limit truncation)
      - Trailing text after JSON
      - Legacy keys (normalised back to the new schema)
    """
    if not text or "impacted_objects" not in text:
        return None

    # Strip markdown code fences (Nova Pro sometimes wraps JSON in them)
    cleaned = re.sub(r'```(?:json)?\s*', '', text)
    cleaned = cleaned.replace('```', '')

    # Find the JSON object using bracket-depth tracking
    for m in re.finditer(r'\{\s*"', cleaned):
        start = m.start()
        depth = 0
        in_str = False
        esc = False
        end = -1
        for i in range(start, len(cleaned)):
            c = cleaned[i]
            if esc:
                esc = False
                continue
            if c == '\\' and in_str:
                esc = True
                continue
            if c == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    end = i
                    break

        candidate = cleaned[start:end + 1] if end != -1 else cleaned[start:]
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict) and "impacted_objects" in parsed:
                return _normalize_to_new_schema(parsed)
        except json.JSONDecodeError:
            repaired = _repair_truncated_json(candidate)
            if repaired and "impacted_objects" in repaired:
                return _normalize_to_new_schema(repaired)
            continue

    return None


def _repair_truncated_json(text: str) -> dict[str, Any] | None:
    """
    Repair JSON that was truncated at a token limit by auto-closing any
    open strings, arrays, and objects. Best-effort — returns None if even
    the patched text won't parse.
    """
    text = text.rstrip()

    # If we ended mid-string, close it
    quote_count = 0
    for i, c in enumerate(text):
        if c == '"' and (i == 0 or text[i-1] != '\\'):
            quote_count += 1
    if quote_count % 2 != 0:
        text += '"'

    # Remove trailing comma or partial value
    text = re.sub(r',\s*$', '', text)
    text = re.sub(r':\s*$', ': null', text)

    # Count open vs close brackets and close them
    open_curly = text.count('{') - text.count('}')
    open_square = text.count('[') - text.count(']')
    text += ']' * max(open_square, 0)
    text += '}' * max(open_curly, 0)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def _extract_nested_dict(text: str, key: str) -> dict[str, Any] | None:
    """Extract the first JSON object for a given key from a text fragment."""
    pos = text.find(key)
    if pos == -1:
        return None
    brace = text.find("{", pos)
    if brace == -1:
        return None
    try:
        decoder = json.JSONDecoder()
        obj, _ = decoder.raw_decode(text, brace)
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# MOCK DATA
# ═══════════════════════════════════════════════════════════════════════════
# Returned by fetch_config_report_from_s3 when the S3 bucket isn't
# configured, so the app can still boot without AWS credentials.

def _mock_config_report() -> dict[str, Any]:
    return {"source": "mock", "objects": [], "integrations": []}
